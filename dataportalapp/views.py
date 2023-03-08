from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
from django.contrib import messages
from .quality import hvaquality, imquality

# Create your views here.
def home(request):
    return render(request, 'home.html')

############################################################################################################################################
######################################################## HUIS VAN ALIJN ####################################################################
############################################################################################################################################
def hva(request):
    return render(request, 'hva.html')

def all(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#000.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Info'
    ws['A1'] = "list of sheet tab codes"
    ws.append(['sheet number', 'quality check'])
    ws.append(['#001', 'instellingsnaam != Het Huis van Alijn (Gent)'])
    ws.append(['#002', 'instellingscode != INST-570'])
    ws.append(['#003', 'objectnummer start niet met: AU-, FO-, 19, 20, DIA-, AF, DB-, RE-, F0, VI-'])
    ws.append(['#004', 'lengte objectnummer niet correct'])
    ws.append(['#005', 'foutief of leeg onderscheidende kenmerken'])
    ws.append(['#006', 'ontbrekende objectnaam'])
    ws.append(['#007', 'ontbrekende titel'])
    ws.append(['#008', 'ontbrekende afbeelding'])
    ws.append(['#009', 'ontbrekende associatie onderwerp'])
    ws.append(['#010', 'vervaardiging plaats niet in associatie onderwerp'])
    ws.append(['#011', 'vervaardiging datering niet in associatie periode'])
    ws.append(['#012', 'vervaardiging datum begin == vervaardiging datum eind'])
    ws.append(['#013', 'vervaardiging datum eind > datum begin'])
    ws.append(['#014', 'geen datering format'])
    ws.append(['#015', 'ontbrekende afmeting'])
    ws.append(['#016', 'objecten afmeting niet in cm'])
    ws.append(['#017', 'digitale collectie afmetingen niet in min, kb, mb of gb'])
    ws.append(['#018', 'documentaire collectie, afmetingen niet in mm'])
    ws.append(['#019', 'ontbrekende rechtenstatus'])
    ws.append(['#020', 'ontbrekende machtigingsstatus'])
    ws.append(['#021', 'ontbrekend formulier'])
    ws.append(['#022', 'publiek domein'])
    ws.append(['#023', 'ontbrekende/foutieve toestand'])
    ws.append(['#024', 'ontbrekende/foutieve verwervingsmethode'])
    ws.append(['#025', 'titel bevat Gent AND NOT associatie.onderwerp = Gent'])
    ws.append(['#026', 'associatie wereldtentoonstelling Brussel, Antwerpen, Gent, Luik, Parijs, Amsterdam, Londen en niet stad als associatie'])
    ws.append(['#t01', 'thesaurus bron afwezig of niet correct'])
    ws.append(['#t02', 'thesaurus term komt meermaals voor'])
    ws.append(['#t03', 'thesaurus externe autoriteit komt meermaals voor'])
    ws.append(['#t04', 'thesaurus foutief wikidatanummer'])
    ws.append(['#t05', 'thesaurus foutief AAT-nummer'])
    ws.append(['#t06', 'thesaurus foutief TGN-nummer'])

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    if df_001.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#001")
        rows = dataframe_to_rows(df_001, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_002.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#002")
        rows = dataframe_to_rows(df_002, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_003.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#003")
        rows = dataframe_to_rows(df_003, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_004.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#004")
        rows = dataframe_to_rows(df_004, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_005.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#005")
        rows = dataframe_to_rows(df_005, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_006.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#006")
        rows = dataframe_to_rows(df_006, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_007.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#007")
        rows = dataframe_to_rows(df_007, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_008.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#008")
        rows = dataframe_to_rows(df_008, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_009.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#009")
        rows = dataframe_to_rows(df_009, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_010.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#010")
        rows = dataframe_to_rows(df_010, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_011.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#011")
        rows = dataframe_to_rows(df_011, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_012.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#012")
        rows = dataframe_to_rows(df_012, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_013.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#013")
        rows = dataframe_to_rows(df_013, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_014.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#014")
        rows = dataframe_to_rows(df_014, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_015.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#015")
        rows = dataframe_to_rows(df_015, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_016.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#016")
        rows = dataframe_to_rows(df_016, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_017.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#017")
        rows = dataframe_to_rows(df_017, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_018.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#018")
        rows = dataframe_to_rows(df_018, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_019.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#019")
        rows = dataframe_to_rows(df_019, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_020.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#020")
        rows = dataframe_to_rows(df_020, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_021.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#021")
        rows = dataframe_to_rows(df_021, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_022.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#022")
        rows = dataframe_to_rows(df_022, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_023.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#023")
        rows = dataframe_to_rows(df_023, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_024.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#024")
        rows = dataframe_to_rows(df_024, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_025.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#025")
        rows = dataframe_to_rows(df_025, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    if df_026.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#026")
        rows = dataframe_to_rows(df_026, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_t01.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#T01")
        rows = dataframe_to_rows(df_t01, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_t02.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#T02")
        rows = dataframe_to_rows(df_t02, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_t03.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#T03")
        rows = dataframe_to_rows(df_t03, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_t04.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#T04")
        rows = dataframe_to_rows(df_t04, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_t05.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#T05")
        rows = dataframe_to_rows(df_t05, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    if df_t06.empty == True:
        print('empty dataframe')
    else:
        ws = wb.create_sheet("#T06")
        rows = dataframe_to_rows(df_t06, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h001(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#001.xlsx"'
    df_001 = hvaquality.h001()
    wb = Workbook()
    if df_001.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#001'
        rows = dataframe_to_rows(df_001, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h002(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#002.xlsx"'
    df_002 = hvaquality.h002()
    wb = Workbook()
    if df_002.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#002'
        rows = dataframe_to_rows(df_002, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h003(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#003.xlsx"'
    df_003 = hvaquality.h003()
    wb = Workbook()
    if df_003.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#003'
        rows = dataframe_to_rows(df_003, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h004(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#004.xlsx"'
    df_004 = hvaquality.h004()
    wb = Workbook()
    if df_004.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#004'
        rows = dataframe_to_rows(df_004, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h005(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#005.xlsx"'
    df_005 = hvaquality.h005()
    wb = Workbook()
    if df_005.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#005'
        rows = dataframe_to_rows(df_005, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h006(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#006.xlsx"'
    df_006 = hvaquality.h006()
    wb = Workbook()
    if df_006.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#006'
        rows = dataframe_to_rows(df_006, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h007(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#007.xlsx"'
    df_007 = hvaquality.h007()
    wb = Workbook()
    if df_007.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#007'
        rows = dataframe_to_rows(df_007, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h008(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#008.xlsx"'
    df_008 = hvaquality.h008()
    wb = Workbook()
    if df_008.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#008'
        rows = dataframe_to_rows(df_008, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h009(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#009.xlsx"'
    df_009 = hvaquality.h009()
    wb = Workbook()
    if df_009.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#009'
        rows = dataframe_to_rows(df_009, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h010(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#010.xlsx"'
    df_010 = hvaquality.h010()
    wb = Workbook()
    if df_010.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#010'
        rows = dataframe_to_rows(df_010, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h011(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#011.xlsx"'
    df_011 = hvaquality.h011()
    wb = Workbook()
    if df_011.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#011'
        rows = dataframe_to_rows(df_011, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h012(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#012.xlsx"'
    df_012 = hvaquality.h012()
    wb = Workbook()
    if df_012.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#012'
        rows = dataframe_to_rows(df_012, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h013(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#013.xlsx"'
    df_013 = hvaquality.h013()
    wb = Workbook()
    if df_013.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#013'
        rows = dataframe_to_rows(df_013, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h014(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#014.xlsx"'
    df_014 = hvaquality.h014()
    wb = Workbook()
    if df_014.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#014'
        rows = dataframe_to_rows(df_014, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h015(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#015.xlsx"'
    df_015 = hvaquality.h015()
    wb = Workbook()
    if df_015.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#015'
        rows = dataframe_to_rows(df_015, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h016(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#016.xlsx"'
    df_016 = hvaquality.h016()
    wb = Workbook()
    if df_016.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#016'
        rows = dataframe_to_rows(df_016, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h017(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#017.xlsx"'
    df_017 = hvaquality.h017()
    wb = Workbook()
    if df_017.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#017'
        rows = dataframe_to_rows(df_017, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h018(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#018.xlsx"'
    df_018 = hvaquality.h018()
    wb = Workbook()
    if df_018.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#018'
        rows = dataframe_to_rows(df_018, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save(response)
        return response

def h019(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#019.xlsx"'
    df_019 = hvaquality.h019()
    wb = Workbook()
    if df_019.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#019'
        rows = dataframe_to_rows(df_019, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h020(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#020.xlsx"'
    df_020 = hvaquality.h020()
    wb = Workbook()
    if df_020.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#020'
        rows = dataframe_to_rows(df_020, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h021(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#021.xlsx"'
    df_021 = hvaquality.h021()
    wb = Workbook()
    if df_021.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#021'
        rows = dataframe_to_rows(df_021, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h022(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#022.xlsx"'
    df_022 = hvaquality.h022()
    wb = Workbook()
    if df_022.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#022'
        rows = dataframe_to_rows(df_022, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h023(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#023.xlsx"'
    df_023 = hvaquality.h023()
    wb = Workbook()
    if df_023.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#023'
        rows = dataframe_to_rows(df_023, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h024(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#024.xlsx"'
    df_024 = hvaquality.h024()
    wb = Workbook()
    if df_024.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#024'
        rows = dataframe_to_rows(df_024, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h025(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#025.xlsx"'
    df_025 = hvaquality.h025()
    wb = Workbook()
    if df_025.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#025'
        rows = dataframe_to_rows(df_025, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h026(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#026.xlsx"'
    df_026 = hvaquality.h026()
    wb = Workbook()
    if df_026.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#026'
        rows = dataframe_to_rows(df_026, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h027(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#027.xlsx"'
    df_027 = hvaquality.h027()
    wb = Workbook()
    if df_027.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#027'
        rows = dataframe_to_rows(df_027, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h028(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#028.xlsx"'
    df_028 = hvaquality.h028()
    wb = Workbook()
    if df_028.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#028'
        rows = dataframe_to_rows(df_028, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h029(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#029.xlsx"'
    df_029 = hvaquality.h029()
    wb = Workbook()
    if df_029.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#029'
        rows = dataframe_to_rows(df_029, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h030(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#030.xlsx"'
    df_030 = hvaquality.h030()
    wb = Workbook()
    if df_030.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#030'
        rows = dataframe_to_rows(df_030, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h031(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#031.xlsx"'
    df_031 = hvaquality.h031()
    wb = Workbook()
    if df_031.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#031'
        rows = dataframe_to_rows(df_031, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h032(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#032.xlsx"'
    df_032 = hvaquality.h032()
    wb = Workbook()
    if df_032.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#032'
        rows = dataframe_to_rows(df_032, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def h033(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#033.xlsx"'
    df_033 = hvaquality.h033()
    wb = Workbook()
    if df_033.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#033'
        rows = dataframe_to_rows(df_033, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ht01(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T01.xlsx"'
    df_t01 = hvaquality.ht01()
    wb = Workbook()
    if df_t01.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#T01'
        rows = dataframe_to_rows(df_t01, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ht02(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T02.xlsx"'
    df_t02 = hvaquality.ht02()
    wb = Workbook()
    if df_t02.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#T02'
        rows = dataframe_to_rows(df_t02, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ht03(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T03.xlsx"'
    df_t03 = hvaquality.ht03()
    wb = Workbook()
    if df_t03.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#T03'
        rows = dataframe_to_rows(df_t03, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ht04(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T04.xlsx"'
    df_t04 = hvaquality.ht04()
    wb = Workbook()
    if df_t04.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#T04'
        rows = dataframe_to_rows(df_t04, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ht05(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T05.xlsx"'
    df_t05 = hvaquality.ht05()
    wb = Workbook()
    if df_t05.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#T05'
        rows = dataframe_to_rows(df_t05, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ht06(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T06.xlsx"'
    df_t06 = hvaquality.ht06()
    wb = Workbook()
    if df_t06.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#T06'
        rows = dataframe_to_rows(df_t06, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def hr01(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R01.xlsx"'
    df_r01 = hvaquality.hr01()
    wb = Workbook()
    if df_r01.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#R01'
        rows = dataframe_to_rows(df_r01, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def hr02(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R02.xlsx"'
    df_r02 = hvaquality.hr02()
    wb = Workbook()
    if df_r02.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#R02'
        rows = dataframe_to_rows(df_r02, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def hr03(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R03.xlsx"'
    df_r03 = hvaquality.hr03()
    wb = Workbook()
    if df_r03.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#R03'
        rows = dataframe_to_rows(df_r03, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def hr04(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R04.xlsx"'
    df_r04 = hvaquality.hr04()
    wb = Workbook()
    if df_r04.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#R04'
        rows = dataframe_to_rows(df_r04, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def hr05(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R05.xlsx"'
    df_r05 = hvaquality.hr05()
    wb = Workbook()
    if df_r05.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#R05'
        rows = dataframe_to_rows(df_r05, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def hr06(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R06.xlsx"'
    df_r06 = hvaquality.hr06()
    wb = Workbook()
    if df_r06.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'hva.html')
    else:
        ws = wb.active
        ws.title = '#R06'
        rows = dataframe_to_rows(df_r06, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

############################################################################################################################################
######################################################## INDUSTRIEMUSEUM ####################################################################
############################################################################################################################################

def im(request):
    return render(request, 'im.html')

def i001(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#001.xlsx"'
    df_001 = imquality.i001()
    wb = Workbook()
    if df_001.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#001'
        rows = dataframe_to_rows(df_001, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i002(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#002.xlsx"'
    df_002 = imquality.i002()
    wb = Workbook()
    if df_002.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#002'
        rows = dataframe_to_rows(df_002, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i003(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#003.xlsx"'
    df_003 = imquality.i003()
    wb = Workbook()
    if df_003.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#003'
        rows = dataframe_to_rows(df_003, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i004(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#004.xlsx"'
    df_004 = imquality.i004()
    wb = Workbook()
    if df_004.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#004'
        rows = dataframe_to_rows(df_004, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i005(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#005.xlsx"'
    df_005 = imquality.i005()
    wb = Workbook()
    if df_005.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#005'
        rows = dataframe_to_rows(df_005, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i006(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#006.xlsx"'
    df_006 = imquality.i006()
    wb = Workbook()
    if df_006.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#006'
        rows = dataframe_to_rows(df_006, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i007(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#007.xlsx"'
    df_007 = imquality.i007()
    wb = Workbook()
    if df_007.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#007'
        rows = dataframe_to_rows(df_007, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i008(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#008.xlsx"'
    df_008 = imquality.i008()
    wb = Workbook()
    if df_008.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#008'
        rows = dataframe_to_rows(df_008, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i009(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#009.xlsx"'
    df_009 = imquality.i009()
    wb = Workbook()
    if df_009.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#009'
        rows = dataframe_to_rows(df_009, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i010(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#010.xlsx"'
    df_010 = imquality.i010()
    wb = Workbook()
    if df_010.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#010'
        rows = dataframe_to_rows(df_010, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i011(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#011.xlsx"'
    df_011 = imquality.i011()
    wb = Workbook()
    if df_011.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#011'
        rows = dataframe_to_rows(df_011, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i012(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#012.xlsx"'
    df_012 = imquality.i012()
    wb = Workbook()
    if df_012.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#012'
        rows = dataframe_to_rows(df_012, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i013(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#013.xlsx"'
    df_013 = imquality.i013()
    wb = Workbook()
    if df_013.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#013'
        rows = dataframe_to_rows(df_013, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i014(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#014.xlsx"'
    df_014 = imquality.i014()
    wb = Workbook()
    if df_014.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#014'
        rows = dataframe_to_rows(df_014, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i015(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#015.xlsx"'
    df_015 = imquality.i015()
    wb = Workbook()
    if df_015.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#015'
        rows = dataframe_to_rows(df_015, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i016(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#016.xlsx"'
    df_016 = imquality.i016()
    wb = Workbook()
    if df_016.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#016'
        rows = dataframe_to_rows(df_016, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i017(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#017.xlsx"'
    df_017 = imquality.i017()
    wb = Workbook()
    if df_017.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#017'
        rows = dataframe_to_rows(df_017, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i018(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#018.xlsx"'
    df_018 = imquality.i018()
    wb = Workbook()
    if df_018.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#018'
        rows = dataframe_to_rows(df_018, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save(response)
        return response

def i019(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#019.xlsx"'
    df_019 = imquality.i019()
    wb = Workbook()
    if df_019.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#019'
        rows = dataframe_to_rows(df_019, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i020(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#020.xlsx"'
    df_020 = imquality.i020()
    wb = Workbook()
    if df_020.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#020'
        rows = dataframe_to_rows(df_020, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i021(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#021.xlsx"'
    df_021 = imquality.i021()
    wb = Workbook()
    if df_021.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#021'
        rows = dataframe_to_rows(df_021, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i022(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#022.xlsx"'
    df_022 = imquality.i022()
    wb = Workbook()
    if df_022.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#022'
        rows = dataframe_to_rows(df_022, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i023(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#023.xlsx"'
    df_023 = imquality.i023()
    wb = Workbook()
    if df_023.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#023'
        rows = dataframe_to_rows(df_023, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i024(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#024.xlsx"'
    df_024 = imquality.i024()
    wb = Workbook()
    if df_024.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#024'
        rows = dataframe_to_rows(df_024, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i025(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#025.xlsx"'
    df_025 = imquality.i025()
    wb = Workbook()
    if df_025.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#025'
        rows = dataframe_to_rows(df_025, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i026(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#026.xlsx"'
    df_026 = imquality.i026()
    wb = Workbook()
    if df_026.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#026'
        rows = dataframe_to_rows(df_026, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def i027(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#027.xlsx"'
    df_027 = imquality.i027()
    wb = Workbook()
    if df_027.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#027'
        rows = dataframe_to_rows(df_027, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def it01(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T01.xlsx"'
    df_t01 = imquality.it01()
    wb = Workbook()
    if df_t01.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#T01'
        rows = dataframe_to_rows(df_t01, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def it02(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T02.xlsx"'
    df_t02 = imquality.it02()
    wb = Workbook()
    if df_t02.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#T02'
        rows = dataframe_to_rows(df_t02, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def it03(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T03.xlsx"'
    df_t03 = imquality.it03()
    wb = Workbook()
    if df_t03.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#T03'
        rows = dataframe_to_rows(df_t03, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def it04(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T04.xlsx"'
    df_t04 = imquality.it04()
    wb = Workbook()
    if df_t04.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#T04'
        rows = dataframe_to_rows(df_t04, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def it05(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T05.xlsx"'
    df_t05 = imquality.it05()
    wb = Workbook()
    if df_t05.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#T05'
        rows = dataframe_to_rows(df_t05, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def it06(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#T06.xlsx"'
    df_t06 = imquality.it06()
    wb = Workbook()
    if df_t06.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#T06'
        rows = dataframe_to_rows(df_t06, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ir01(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R01.xlsx"'
    df_r01 = imquality.ir01()
    wb = Workbook()
    if df_r01.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#R01'
        rows = dataframe_to_rows(df_r01, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ir02(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R02.xlsx"'
    df_r02 = imquality.ir02()
    wb = Workbook()
    if df_r02.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#R02'
        rows = dataframe_to_rows(df_r02, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ir03(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R03.xlsx"'
    df_r03 = imquality.ir03()
    wb = Workbook()
    if df_r03.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#R03'
        rows = dataframe_to_rows(df_r03, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ir04(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R04.xlsx"'
    df_r04 = imquality.ir04()
    wb = Workbook()
    if df_r04.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#R04'
        rows = dataframe_to_rows(df_r04, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response

def ir05(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="#R05.xlsx"'
    df_r05 = imquality.ir05()
    wb = Workbook()
    if df_r05.empty == True:
        messages.success(request, 'Congrats! Empty list :)')
        return render(request, 'im.html')
    else:
        ws = wb.active
        ws.title = '#R05'
        rows = dataframe_to_rows(df_r05, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(response)
    return response
